import win32com.client
import re
import pandas as pd
from openpyxl import load_workbook, Workbook
import os
from datetime import datetime, timedelta, time, timezone
import time as time_module
import pythoncom
import sys
import logging

# === Определяем базовую папку ===
def get_base_path():
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    else:
        return os.path.dirname(os.path.abspath(__file__))

BASE_PATH = get_base_path()

# === ПУТЬ К ТАБЛИЦЕ С ДАННЫМИ ===
TABLE_FILE = r"C:\Users\RULegkiiAn.NESTLESOFT\OneDrive - Nestle Russia & Eurasia\RU DO Warehouses - оборот поддонов\Екатеринбург - учет оборота поддонов.xlsx"

# === Логирование ===
log_file = os.path.join(BASE_PATH, "script.log")
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(log_file, encoding='utf-8'),
        logging.StreamHandler(sys.stdout)
    ]
)

SUPPORT_CONTACT = "andrei.legkii@nestle.ru"

# === Настройки ===
OUTLOOK_FOLDER = "Inbox"
SEARCH_SUBJECT = "Возврат поддонов из сетей"
EXCEL_FILE = os.path.join(BASE_PATH, "возврат_поддонов.xlsx")
PROCESSED_IDS_FILE = os.path.join(BASE_PATH, "processed_ids.txt")
SHEET_NAME = "Данные"
WRITE_MODE = "horizontal"

# === ПОЛУЧАТЕЛИ И ОТПРАВИТЕЛЬ НАПОМИНАНИЙ ===
REMINDER_RECIPIENTS = {
    "Х5": ["dma@line7.ru", "slon07@line7.ru", "rudcekb@nestlesoft.net"],
    "Тандер": ["rudcekb@nestlesoft.net"],
    "дистры": ["rudcekb@nestlesoft.net"]
}
SENDER_EMAIL = "andrei.legkii@nestle.ru"  # Замените на нужный вам адрес

# === ФЛАГ ТЕСТИРОВАНИЯ (поставьте False для продакшена) ===
TEST_MODE = False  # True — для тестирования, False — для реальной работы
TEST_HOUR = 16    # Час для теста (12 или 14)
TEST_MINUTE = 2   # Минута для теста

# === Глобальные переменные ===
sent_reminders = set()
_processed_ids = set()


def set_console_title(title):
    try:
        import ctypes
        ctypes.windll.kernel32.SetConsoleTitleW(title)
    except Exception as e:
        logging.debug(f"Не удалось установить заголовок консоли: {e}")


def load_processed_ids():
    global _processed_ids
    if not os.path.exists(PROCESSED_IDS_FILE):
        return set()
    try:
        with open(PROCESSED_IDS_FILE, "r", encoding="utf-8") as f:
            _processed_ids = set(line.strip() for line in f if line.strip())
    except Exception as e:
        logging.error(f"Ошибка загрузки processed_ids: {e}")
    return _processed_ids


def save_processed_ids(ids):
    try:
        with open(PROCESSED_IDS_FILE, "w", encoding="utf-8") as f:
            for item_id in ids:
                f.write(item_id + "\n")
    except Exception as e:
        logging.error(f"Ошибка сохранения processed_ids: {e}")


def is_email_in_excel(entry_id):
    if not os.path.exists(EXCEL_FILE):
        return False
    try:
        df = pd.read_excel(EXCEL_FILE, sheet_name=SHEET_NAME)
        if "EntryID" in df.columns and entry_id in df["EntryID"].values:
            return True
    except Exception as e:
        logging.debug(f"Ошибка проверки Excel: {e}")
    return False


def parse_email(body, received_time):
    try:
        lines = body.splitlines()
        default_date_str = received_time.strftime("%Y-%m-%d %H:%M")
        data = {
            "Дата письма": default_date_str,
            "Дата возврата из письма": None,
            "Сеть": "",
            "РЦ": "",
            "Тягач": "",
            "Прицеп": "",
            "ФИО водителя": "",
            "Паспорт": "",
            "Номер ВУ": "",
            "Телефон": "",
            "ИНН": "",
            "Доп. информация": "",
            "EntryID": ""
        }

        for line in lines:
            line = line.strip()
            if not line:
                continue

            if line.startswith("Дата") and "возврат" in line.lower():
                parts = line.split(" ")
                if len(parts) >= 2:
                    date_part = parts[1]
                    try:
                        dt_obj = datetime.strptime(date_part, "%d.%m.%Y").date()
                        data["Дата возврата из письма"] = dt_obj
                        combined_dt = dt_obj.strftime("%Y-%m-%d") + " " + received_time.strftime("%H:%M")
                        data["Дата письма"] = combined_dt
                        logging.debug(f"📅 Найдена дата возврата: {date_part}")
                    except ValueError:
                        logging.warning(f"⚠️ Не удалось распарсить дату: '{line}'")

            elif line.startswith("Сеть"):
                parts = [p.strip() for p in line.split("|")]
                if len(parts) >= 2:
                    data["Сеть"] = parts[1]
                if len(parts) >= 3:
                    data["РЦ"] = parts[2].strip()
            elif line.startswith("Тягач"):
                data["Тягач"] = line.split(":", 1)[1].strip() if ":" in line else ""
            elif line.startswith("Прицеп"):
                data["Прицеп"] = line.split(":", 1)[1].strip() if ":" in line else ""
            elif line.startswith("Ф.И.О. водителя"):
                data["ФИО водителя"] = line.split(":", 1)[1].strip() if ":" in line else ""
            elif line.startswith("Паспорт"):
                data["Паспорт"] = line.split(":", 1)[1].strip() if ":" in line else ""
            elif line.startswith("Номер ВУ"):
                data["Номер ВУ"] = line.split(":", 1)[1].strip() if ":" in line else ""
            elif line.startswith("Телефон"):
                data["Телефон"] = line.split(":", 1)[1].strip() if ":" in line else ""
            elif line.startswith("ИНН"):
                data["ИНН"] = line.split(":", 1)[1].strip() if ":" in line else ""
            elif line.startswith("Дополнительная информация"):
                data["Доп. информация"] = line.split(":", 1)[1].strip() if ":" in line else ""

        return data

    except Exception as e:
        logging.error(f"Ошибка парсинга письма: {e}")
        return None


def send_email(subject, body, to):
    try:
        outlook_app = win32com.client.Dispatch("Outlook.Application")

        mail = outlook_app.CreateItem(0)

        # === НАСТРОЙКА ОТПРАВИТЕЛЯ ===
        accounts = outlook_app.Session.Accounts
        sender_account = None
        for account in accounts:
            if account.SmtpAddress == SENDER_EMAIL:
                sender_account = account
                break
        if sender_account:
            mail.SendUsingAccount = sender_account
            logging.debug(f"📧 Используем аккаунт: {SENDER_EMAIL}")
        else:
            logging.warning(f"⚠️ Аккаунт с адресом {SENDER_EMAIL} не найден в Outlook. Используется основной.")
        # === /НАСТРОЙКА ОТПРАВИТЕЛЯ ===

        mail.Subject = subject
        if isinstance(to, list):
            mail.To = ";".join(to)
        else:
            mail.To = to
        mail.Body = body
        mail.Send()
        logging.info(f"✅ Отправлено уведомление: {subject} -> {to} от {SENDER_EMAIL}")
        del mail
        del outlook_app
    except Exception as e:
        logging.error(f"❌ Ошибка отправки email: {e}")


def update_table_row(data, target_date, target_rc):
    logging.debug(f"🔍 Обновление таблицы: дата={target_date}, РЦ='{target_rc}'")
    try:
        if not os.path.exists(TABLE_FILE):
            logging.warning(f"Файл таблицы не найден: {TABLE_FILE}")
            return False

        book = load_workbook(TABLE_FILE, keep_vba=False)
        if "приход" not in book.sheetnames:
            logging.error("Лист 'приход' не найден!")
            return False

        sheet = book["приход"]
        headers = [str(cell.value).strip() if cell.value else "" for cell in sheet[1]]

        date_col_idx = rc_col_idx = driver_col_idx = tractor_col_idx = None
        for i, h in enumerate(headers):
            if h.lower() == "дата":
                date_col_idx = i + 1
            elif "рц" in h.lower() and "(выберите из списка)" in h.lower():
                rc_col_idx = i + 1
            elif "водитель" in h.lower() and "фамилия" in h.lower():
                driver_col_idx = i + 1
            elif h.lower() == "номер ам":
                tractor_col_idx = i + 1

        missing = [col for col, name in [("дата", date_col_idx), ("РЦ", rc_col_idx), ("водитель", driver_col_idx), ("номер ам", tractor_col_idx)] if name is None]
        if missing:
            logging.error(f"Не найдены столбцы: {', '.join(missing)}")
            return False

        found_row = None
        for row_num in range(2, sheet.max_row + 1):
            cell_date = sheet.cell(row=row_num, column=date_col_idx).value
            if cell_date is None:
                continue
            try:
                if isinstance(cell_date, str):
                    parsed = datetime.strptime(cell_date, "%d.%m.%Y").date() if "." in cell_date else datetime.strptime(cell_date, "%Y-%m-%d").date()
                else:
                    parsed = cell_date.date()
            except:
                continue
            if parsed == target_date and str(sheet.cell(row=row_num, column=rc_col_idx).value).strip() == target_rc:
                found_row = row_num
                break

        if not found_row:
            logging.info("Строка не найдена — пропуск обновления")
            return False

        updated = False
        for col_idx, key in [(driver_col_idx, "ФИО водителя"), (tractor_col_idx, "Тягач")]:
            val = data.get(key, "").strip()
            if val:
                cell = sheet.cell(row=found_row, column=col_idx)
                if not cell.value or str(cell.value).strip().lower() in ("", "nan"):
                    cell.value = val
                    updated = True
                    logging.info(f"✅ Обновлено: {key} = {val}")

        if updated:
            book.save(TABLE_FILE)
            logging.info("💾 Таблица сохранена")
        return True

    except Exception as e:
        logging.error(f"❌ Ошибка обновления таблицы: {e}")
        return False


# 🔥 ОСНОВНАЯ ФУНКЦИЯ С ПОЛНЫМ ЛОГИРОВАНИЕМ И ТЕСТОВЫМ РЕЖИМОМ
def check_reminders_from_table():
    global sent_reminders

    try:
        logging.info("🔍 Запуск проверки напоминаний...")
        if not os.path.exists(TABLE_FILE):
            logging.error("❌ Файл таблицы не найден — пропуск проверки напоминаний")
            return

        # === ОПРЕДЕЛЕНИЕ ВРЕМЕНИ (реальное или тестовое) ===
        if TEST_MODE:
            now_msk = datetime.now().replace(hour=TEST_HOUR, minute=TEST_MINUTE, second=0, microsecond=0)
            logging.info(f"🧪 ТЕСТОВЫЙ РЕЖИМ: используем время {now_msk.strftime('%H:%M:%S')}")
        else:
            moscow_tz = timezone(timedelta(hours=3))
            now_msk = datetime.now(moscow_tz)
            logging.info(f"🌍 РЕАЛЬНОЕ ВРЕМЯ: {now_msk.strftime('%H:%M:%S')} по МСК")

        today_msk = now_msk.date()
        current_time_msk = now_msk.time()

        book = load_workbook(TABLE_FILE, read_only=True, data_only=True)
        if "приход" not in book.sheetnames:
            logging.error("❌ Лист 'приход' не найден в таблице")
            book.close()
            return

        sheet = book["приход"]
        headers = [str(cell.value).strip() if cell.value else "" for cell in sheet[1]]

        logging.debug(f"📋 Заголовки столбцов: {headers}")

        col_map = {}
        for i, h in enumerate(headers):
            if h.lower() == "дата":
                col_map["дата"] = i
            elif "рц" in h.lower() and "(выберите из списка)" in h.lower():
                col_map["рц"] = i
            elif "поставщик" in h.lower() and "(выберите из списка)" in h.lower():
                col_map["поставщик"] = i
            elif "водитель" in h.lower() and "фамилия" in h.lower():
                col_map["водитель"] = i
            elif h.lower() == "номер ам":
                col_map["тягач"] = i

        required = ["дата", "рц", "поставщик", "водитель", "тягач"]
        if not all(k in col_map for k in required):
            missing = [k for k in required if k not in col_map]
            logging.error(f"❌ Не найдены столбцы: {', '.join(missing)}")
            book.close()
            return

        logging.info("🔍 Начинаю проверку всех строк таблицы для напоминаний...")

        row_index = 1
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_index += 1
            cell_date = row[col_map["дата"]]
            if cell_date is None:
                continue

            try:
                parsed_date = datetime.strptime(cell_date, "%d.%m.%Y").date() if isinstance(cell_date, str) and "." in cell_date else \
                              datetime.strptime(cell_date, "%Y-%m-%d").date() if isinstance(cell_date, str) else cell_date.date()
            except:
                logging.debug(f"⚠️ Строка {row_index}: не удалось распарсить дату '{cell_date}' — пропуск")
                continue

            rc = str(row[col_map["рц"]]).strip()
            raw_supplier = row[col_map["поставщик"]]
            supplier = str(raw_supplier).strip() if raw_supplier is not None else ""
            supplier_lower = supplier.lower()

            # Очистка от невидимых символов
            supplier = ''.join(c for c in supplier if c.isprintable())

            # Проверка, заполнены ли водитель и тягач
            raw_driver = row[col_map["водитель"]]
            raw_tractor = row[col_map["тягач"]]
            has_driver = bool(raw_driver and str(raw_driver).strip().lower() not in ("", "nan", "none"))
            has_tractor = bool(raw_tractor and str(raw_tractor).strip().lower() not in ("", "nan", "none"))

            logging.debug(f"📋 Строка {row_index}: дата={parsed_date}, РЦ={rc}, поставщик='{supplier}', водитель={'да' if has_driver else 'нет'}, тягач={'да' if has_tractor else 'нет'}")

            # === X5 и Дистры: напоминание в 12:00 по МСК в день возврата ===
            if ("x5" in supplier_lower or "х5" in supplier_lower or "дистр" in supplier_lower) and today_msk == parsed_date:
                logging.info(f"✅ X5/Дистры: строка {row_index} подходит по дате и поставщику")
                if TEST_MODE and current_time_msk.hour == TEST_HOUR or (not TEST_MODE and time(12, 0) <= current_time_msk < time(12, 1)):
                    logging.info("⏰ Время 12:00 (или тестовое) — проверяю, нужно ли отправлять напоминание")
                    if not (has_driver and has_tractor):
                        key = (parsed_date.isoformat(), rc, "x5_distry_12h")
                        if key not in sent_reminders:
                            net_name = "X5" if "x5" in supplier_lower or "х5" in supplier_lower else "Дистры"
                            subject = f"📅 Напоминание ({net_name}): предоставьте данные водителя на РЦ {rc}"
                            body = (
                                f"Дата возврата: {parsed_date.strftime('%d.%m.%Y')}\n"
                                f"Поставщик: {supplier}\n"
                                f"РЦ: {rc}\n\n"
                                f"Напоминаем предоставить данные для оформления пропуска.\n"
                                f"[Автоматическое уведомление]"
                            )
                            # === ВЫБОР ПОЛУЧАТЕЛЕЙ ПО СЕТИ ===
                            recipients = REMINDER_RECIPIENTS.get("x5" if "x5" in supplier_lower or "х5" in supplier_lower else "дистры")
                            send_email(subject, body, recipients)
                            sent_reminders.add(key)
                            logging.info(f"✅ ✉️ ОТПРАВЛЕНО напоминание для {net_name} на {recipients} от {SENDER_EMAIL}")
                        else:
                            logging.info(f"ℹ️ Напоминание для {supplier} уже отправлялось сегодня")
                    else:
                        logging.info(f"ℹ️ Для {supplier} данные уже заполнены — напоминание не нужно")
                else:
                    logging.info(f"⏳ Время {current_time_msk} — не 12:00 (или не тестовое). Пропуск.")

            # === Тандер: напоминание в 14:00 по МСК накануне ===
            elif "тандер" in supplier_lower:
                weekday = parsed_date.weekday()
                if weekday in (5, 6, 0):  # Сб, Вс, Пн → напоминание в Пт
                    days_to_fri = (weekday - 4) % 7
                    if days_to_fri == 0:
                        days_to_fri = 7
                    reminder_date = parsed_date - timedelta(days=days_to_fri)
                    logging.debug(f"📅 Тандер: возврат {parsed_date} (Пн/Сб/Вс) → напоминание в {reminder_date} (Пт)")
                else:
                    reminder_date = parsed_date - timedelta(days=1)
                    logging.debug(f"📅 Тандер: возврат {parsed_date} → напоминание накануне: {reminder_date}")

                if today_msk == reminder_date:
                    logging.info(f"✅ Сегодня день напоминания для Тандера (возврат {parsed_date}, напоминание {reminder_date})")
                    if TEST_MODE and current_time_msk.hour == TEST_HOUR or (not TEST_MODE and time(14, 0) <= current_time_msk < time(14, 1)):
                        logging.info("⏰ Сейчас 14:00 (или тестовое) — проверяю необходимость отправки напоминания")
                        if not (has_driver and has_tractor):
                            key = (reminder_date.isoformat(), rc, "tander_14h")
                            if key not in sent_reminders:
                                subject = f"ТАНДЕР: срочно предоставьте данные водителя на РЦ {rc}"
                                body = (
                                    f"Дата возврата: {parsed_date.strftime('%d.%m.%Y')}\n"
                                    f"Поставщик: {supplier}\n"
                                    f"РЦ: {rc}\n\n"
                                    f"Данные водителя или тягача отсутствуют в таблице учёта.\n"
                                    f"[Автоматическое уведомление]"
                                )
                                # === ВЫБОР ПОЛУЧАТЕЛЕЙ ПО СЕТИ ===
                                recipients = REMINDER_RECIPIENTS["тандер"]
                                send_email(subject, body, recipients)
                                sent_reminders.add(key)
                                logging.info(f"✅ ✉️ ОТПРАВЛЕНО напоминание для ТАНДЕР на {recipients} от {SENDER_EMAIL}")
                            else:
                                logging.info("ℹ️ Напоминание для Тандера уже отправлялось сегодня")
                        else:
                            logging.info("ℹ️ Для Тандера данные уже заполнены — напоминание не требуется")
                    else:
                        logging.info(f"⏳ Время {current_time_msk} — не в окне 14:00 (или не тестовое). Пропуск.")
                else:
                    logging.debug(f"📆 Напоминание для Тандера не сегодня (ожидается {reminder_date})")

        book.close()
        logging.info("✅ Завершена проверка напоминаний по таблице")

    except Exception as e:
        logging.error(f"❌ КРИТИЧЕСКАЯ ОШИБКА в check_reminders_from_table: {e}", exc_info=True)


# === Запись в Excel ===
def write_vertical_to_excel(data, sheet_name, excel_file):
    try:
        if not os.path.exists(excel_file):
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            ws.append(["Ключ", "Значение", "EntryID"])
            wb.save(excel_file)
            logging.info(f"✅ Создан вертикальный Excel: {excel_file}")
        else:
            book = load_workbook(excel_file)
            if sheet_name not in book.sheetnames:
                ws = book.create_sheet(sheet_name)
                ws.append(["Ключ", "Значение", "EntryID"])
            else:
                ws = book[sheet_name]
            startrow = ws.max_row + 1
            ws.cell(row=startrow, column=1, value=f"=== Письмо от {data['Дата письма']} ===")
            ws.cell(row=startrow, column=3, value=data["EntryID"])
            startrow += 1
            for k, v in data.items():
                if k == "EntryID": continue
                ws.cell(row=startrow, column=1, value=k)
                ws.cell(row=startrow, column=2, value=v)
                ws.cell(row=startrow, column=3, value=data["EntryID"])
                startrow += 1
            book.save(excel_file)
            logging.info("✅ Записано вертикально")
    except Exception as e:
        logging.error(f"❌ Ошибка вертикальной записи: {e}")


def write_horizontal_to_excel(data, sheet_name, excel_file):
    try:
        df_new = pd.DataFrame([data])
        if not os.path.exists(EXCEL_FILE):
            df_new.to_excel(EXCEL_FILE, sheet_name=sheet_name, index=False)
            logging.info(f"✅ Создан горизонтальный Excel: {excel_file}")
        else:
            book = load_workbook(EXCEL_FILE)
            if sheet_name not in book.sheetnames:
                with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl", mode="a") as writer:
                    df_new.to_excel(writer, sheet_name=sheet_name, index=False)
            else:
                startrow = book[SHEET_NAME].max_row
                with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                    df_new.to_excel(writer, sheet_name=SHEET_NAME, index=False, header=False, startrow=startrow)
            logging.info("✅ Записано горизонтально")
    except Exception as e:
        logging.error(f"❌ Ошибка горизонтальной записи: {e}")


def handle_mail(item, processed_ids):
    global _processed_ids
    try:
        entry_id = item.EntryID
        subject = item.Subject
        received_time = item.ReceivedTime
        logging.info(f"🔍 Обработка: {subject} | ID: {entry_id}")

        if entry_id in processed_ids or is_email_in_excel(entry_id):
            processed_ids.add(entry_id)
            save_processed_ids(processed_ids)
            return

        if not subject or SEARCH_SUBJECT.lower() not in str(subject).lower():
            return

        body = item.Body
        del item

        data = parse_email(body, received_time)
        if not data:
            logging.warning("⚠️ Не удалось распарсить письмо")
            processed_ids.add(entry_id)
            save_processed_ids(processed_ids)
            return

        data["EntryID"] = entry_id
        logging.info(f"✅ Извлечено: {data}")

        target_date = data.get("Дата возврата из письма")
        if not target_date:
            try:
                target_date = datetime.strptime(data.get("Дата письма", "")[:10], "%Y-%m-%d").date()
            except:
                target_date = None

        target_rc = data.get("РЦ", "").strip()

        if target_date and target_rc:
            update_table_row(data, target_date, target_rc)

        if WRITE_MODE == "vertical":
            write_vertical_to_excel(data, SHEET_NAME, EXCEL_FILE)
        else:
            write_horizontal_to_excel(data, SHEET_NAME, EXCEL_FILE)

        # ❌ НЕТ вызова напоминаний здесь — они в отдельной функции

        processed_ids.add(entry_id)
        save_processed_ids(processed_ids)

    except Exception as e:
        logging.error(f"❌ Ошибка обработки письма: {e}")
    finally:
        if 'item' in locals():
            del item


def monitor_inbox():
    pythoncom.CoInitialize()
    outlook = namespace = inbox = folder = None

    processed_ids = load_processed_ids()
    logging.info(f"Загружено {len(processed_ids)} обработанных писем.")

    try:
        outlook = win32com.client.Dispatch("Outlook.Application")
        namespace = outlook.GetNamespace("MAPI")
        inbox = namespace.GetDefaultFolder(6)
        folder = inbox.Folders[OUTLOOK_FOLDER] if OUTLOOK_FOLDER.lower() != "inbox" else inbox

        logging.info(f"✅ Мониторинг запущен. Режим: {WRITE_MODE.upper()}. Папка: {OUTLOOK_FOLDER}")
        time_module.sleep(5)

        while True:
            try:
                min_date = datetime.today().date() - timedelta(days=7)
                messages = folder.Items
                messages.Sort("[ReceivedTime]", True)

                msg_list = []
                for msg in messages:
                    try:
                        if getattr(msg, 'Class', None) != 43:
                            continue
                        if not getattr(msg, 'ReceivedTime', None):
                            continue
                        if msg.ReceivedTime.date() < min_date:
                            break
                        msg_list.append(msg)
                    except Exception as e:
                        logging.error(f"Ошибка при анализе письма: {e}")

                for msg in msg_list:
                    handle_mail(msg, processed_ids)
                    del msg

                # ✅ Отдельная проверка напоминаний
                check_reminders_from_table()

            except Exception as e:
                logging.error(f"❌ Ошибка мониторинга: {e}")

            logging.info("⏳ Ждем 60 секунд...\n")
            time_module.sleep(60)

    except Exception as e:
        logging.error(f"❌ Ошибка инициализации Outlook: {e}")
    finally:
        for obj in [folder, inbox, namespace, outlook]:
            if obj:
                del obj
        pythoncom.CoUninitialize()


if __name__ == "__main__":
    try:
        set_console_title("📦 Система учета возврата поддонов")
        logging.info("=" * 50)
        logging.info("  📦 Система учета возврата поддонов")
        logging.info(f"  📞 Поддержка: {SUPPORT_CONTACT}")
        logging.info("=" * 50)
        logging.info("")
        monitor_inbox()
    except Exception as e:
        logging.error(f"❌ Критическая ошибка: {e}")
    finally:
        input("\nНажмите Enter для закрытия...")
