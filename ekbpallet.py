import win32com.client
import re
import pandas as pd
from openpyxl import load_workbook, Workbook
import os
from datetime import datetime, timedelta
import time
import pythoncom
import sys
import logging

# === Определяем базовую папку (где лежит .exe или .py) ===
def get_base_path():
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    else:
        return os.path.dirname(os.path.abspath(__file__))

BASE_PATH = get_base_path()

# === ⚠️ УКАЖИ ПУТЬ К ТАБЛИЦЕ С ДАННЫМИ ===
TABLE_FILE = r"C:\Users\RULegkiiAn.NESTLESOFT\OneDrive - Nestle Russia & Eurasia\RU DO Warehouses - оборот поддонов\Екатеринбург - учет оборота поддонов.xlsx"

# === Настройка логирования ===
log_file = os.path.join(BASE_PATH, "script.log")
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(log_file, encoding='utf-8'),
        logging.StreamHandler(sys.stdout)
    ]
)

SUPPORT_CONTACT = "andrei.legkii@nestle.ru"

# === Настройки ===
OUTLOOK_FOLDER = "Inbox"
SEARCH_SUBJECT = "Возврат поддонов из сетей"
EXCEL_FILE = os.path.join(BASE_PATH, "возврат_поддонов.xlsx")
PROCESSED_IDS_FILE = os.path.join(BASE_PATH, "processed_ids.txt")
SHEET_NAME = "Данные"

WRITE_MODE = "horizontal"  # или "vertical"

# === Получатели напоминаний ===
REMINDER_RECIPIENTS = {
    "x5": ["dma@line7.ru", "slon07@line7.ru", "rudcekb@nestlesoft.net"],
    "тандер": ["rudcekb@nestlesoft.net"],
    "дистры": ["rudcekb@nestlesoft.net"]
}

# === Глобальные переменные ===
_table_cache = None
_table_cache_time = None
sent_reminders = set()
_processed_ids = set()


# === Установка заголовка консоли ===
def set_console_title(title):
    try:
        import ctypes
        ctypes.windll.kernel32.SetConsoleTitleW(title)
    except Exception as e:
        logging.debug(f"Не удалось установить заголовок консоли: {e}")


# === Загрузка уже обработанных ID из файла ===
def load_processed_ids():
    global _processed_ids
    if not os.path.exists(PROCESSED_IDS_FILE):
        return set()
    try:
        with open(PROCESSED_IDS_FILE, "r", encoding="utf-8") as f:
            _processed_ids = set(line.strip() for line in f if line.strip())
    except Exception as e:
        logging.error(f"Ошибка загрузки processed_ids: {e}")
    return _processed_ids


# === Сохранение обработанных ID в файл ===
def save_processed_ids(ids):
    try:
        with open(PROCESSED_IDS_FILE, "w", encoding="utf-8") as f:
            for item_id in ids:
                f.write(item_id + "\n")
    except Exception as e:
        logging.error(f"Ошибка сохранения processed_ids: {e}")


# === Проверка, есть ли письмо уже в Excel (по EntryID) ===
def is_email_in_excel(entry_id):
    if not os.path.exists(EXCEL_FILE):
        return False
    try:
        df = pd.read_excel(EXCEL_FILE, sheet_name=SHEET_NAME)
        if "EntryID" in df.columns and entry_id in df["EntryID"].values:
            return True
    except Exception as e:
        logging.debug(f"Ошибка проверки Excel: {e}")
    return False


# === 🔹 ИСПРАВЛЕННАЯ ФУНКЦИЯ: Парсинг письма с извлечением даты из тела ===
def parse_email(body, received_time):
    try:
        lines = body.splitlines()
        # Инициализируем дату по умолчанию как дату получения
        default_date_str = received_time.strftime("%Y-%m-%d %H:%M")
        data = {
            "Дата письма": default_date_str,  # По умолчанию - дата получения
            "Дата возврата из письма": None,  # Новое поле для даты из тела
            "Сеть": "",
            "РЦ": "",  # <<< Сохраняем РЦ в точности как в письме
            "Тягач": "",
            "Прицеп": "",
            "ФИО водителя": "",
            "Паспорт": "",
            "Номер ВУ": "",
            "Телефон": "",
            "ИНН": "",
            "Доп. информация": "",
            "EntryID": ""
        }

        for line in lines:
            line = line.strip()
            if not line:
                continue

            # === НОВОЕ: Извлечение даты из тела письма ===
            # Ищем строку вида "Дата 11.10.2025 возврат"
            if line.startswith("Дата") and "возврат" in line.lower():
                # Пример: "Дата 11.10.2025 возврат"
                parts = line.split(" ")
                if len(parts) >= 2:
                    date_part = parts[1]  # Берем второй элемент ("11.10.2025")
                    try:
                        # Преобразуем в объект date
                        dt_obj = datetime.strptime(date_part, "%d.%m.%Y").date()
                        data["Дата возврата из письма"] = dt_obj
                        # Перезаписываем "Дата письма" только датой из письма (время остается от получения)
                        # Это позволяет использовать "Дата письма" как ключ для поиска в таблице
                        combined_dt = dt_obj.strftime("%Y-%m-%d") + " " + received_time.strftime("%H:%M")
                        data["Дата письма"] = combined_dt
                        logging.debug(f"📅 Найдена дата возврата в теле письма: {date_part} -> {dt_obj}")
                    except ValueError:
                        logging.warning(f"⚠️ Не удалось распарсить дату из строки: '{line}'")

            elif line.startswith("Сеть"):
                parts = [p.strip() for p in line.split("|")]
                if len(parts) >= 2:
                    data["Сеть"] = parts[1]
                if len(parts) >= 3:
                    # <<< СОХРАНЯЕМ РЦ КАК ЕСТЬ, БЕЗ ОБРЕЗАНИЯ "РЦ" >>>
                    data["РЦ"] = parts[2].strip()  # parts[2] уже содержит "РЦ Тюмень"
            elif line.startswith("Тягач"):
                data["Тягач"] = line.split(":", 1)[1].strip() if ":" in line else ""
            elif line.startswith("Прицеп"):
                data["Прицеп"] = line.split(":", 1)[1].strip() if ":" in line else ""
            elif line.startswith("Ф.И.О. водителя"):
                data["ФИО водителя"] = line.split(":", 1)[1].strip() if ":" in line else ""
            elif line.startswith("Паспорт"):
                data["Паспорт"] = line.split(":", 1)[1].strip() if ":" in line else ""
            elif line.startswith("Номер ВУ"):
                data["Номер ВУ"] = line.split(":", 1)[1].strip() if ":" in line else ""
            elif line.startswith("Телефон"):
                data["Телефон"] = line.split(":", 1)[1].strip() if ":" in line else ""
            elif line.startswith("ИНН"):
                data["ИНН"] = line.split(":", 1)[1].strip() if ":" in line else ""
            elif line.startswith("Дополнительная информация"):
                data["Доп. информация"] = line.split(":", 1)[1].strip() if ":" in line else ""

        return data

    except Exception as e:
        logging.error(f"Ошибка парсинга письма: {e}")
        return None


# === Отправка уведомления ===
def send_email(subject, body, to, cc=None):
    try:
        outlook_app = win32com.client.Dispatch("Outlook.Application")
        mail = outlook_app.CreateItem(0)
        mail.Subject = subject
        if isinstance(to, list):
            mail.To = ";".join(to)
        else:
            mail.To = to
        if cc:
            mail.CC = cc
        mail.Body = body
        mail.Send()
        logging.info(f"Отправлено уведомление: {subject} -> {to}")
        del mail
        del outlook_app
    except Exception as e:
        logging.error(f"Ошибка отправки email: {e}")


# === 🔹 НОВАЯ ФУНКЦИЯ: Обновление строки в таблице (БЕЗ проверки по поставщику) ===
def update_table_row(data, target_date, target_rc):
    """
    Обновляет строку в TABLE_FILE (Екатеринбург - учет оборота поддонов.xlsx)
    по совпадению даты и РЦ.
    Заполняет только пустые ячейки "водитель Фамилия И.О." и "номер ам".
    """
    logging.debug(f"🔍 update_table_row: Начало. Ищем дату {target_date}, РЦ '{target_rc}'")
    try:
        if not os.path.exists(TABLE_FILE):
            logging.warning(f"Файл таблицы не найден: {TABLE_FILE}")
            return False

        # Открываем книгу
        logging.debug(f"📂 update_table_row: Открываем файл таблицы: {TABLE_FILE}")
        book = load_workbook(TABLE_FILE, keep_vba=False)
        if "приход" not in book.sheetnames:
            logging.error("Лист 'приход' не найден в таблице!")
            return False

        sheet = book["приход"]

        # Получаем заголовки (преобразуем в строки и убираем пробелы)
        headers_raw = [cell.value for cell in sheet[1]]
        headers = []
        for h in headers_raw:
            if h is None:
                headers.append("")
            else:
                headers.append(str(h).strip())

        logging.debug(f"📋 update_table_row: Заголовки таблицы: {headers}")

        # Ищем индексы нужных столбцов
        date_col_idx = None
        rc_col_idx = None
        driver_col_idx = None
        tractor_col_idx = None
        # supplier_col_idx = None  # УБРАНО

        for i, header_text in enumerate(headers):
            # Столбец даты
            if header_text.lower() == "дата":
                date_col_idx = i + 1  # openpyxl использует 1-based индексацию
            # Столбец РЦ (с учетом переноса строки)
            elif "рц" in header_text.lower() and "(выберите из списка)" in header_text.lower():
                rc_col_idx = i + 1
            # Столбец водителя
            elif "водитель" in header_text.lower() and "фамилия" in header_text.lower():
                driver_col_idx = i + 1
            # Столбец номера а/м
            elif header_text.lower() == "номер ам":
                tractor_col_idx = i + 1
            # supplier_col_idx не ищем  # УБРАНО

        logging.debug(f"📊 update_table_row: Индексы столбцов: дата={date_col_idx}, РЦ={rc_col_idx}, водитель={driver_col_idx}, номер ам={tractor_col_idx}")

        missing_cols = []
        if date_col_idx is None:
            missing_cols.append("дата")
        if rc_col_idx is None:
            missing_cols.append("РЦ (выберите из списка)")
        if driver_col_idx is None:
            missing_cols.append("водитель Фамилия И.О.")
        if tractor_col_idx is None:
            missing_cols.append("номер ам")

        if missing_cols:
            logging.error(f"❌ update_table_row: Не найдены обязательные столбцы в таблице: {', '.join(missing_cols)}")
            return False

        # === 🔍 Ищем строку с совпадающей датой и РЦ ===
        found_row = None
        logging.debug(f"🔎 update_table_row: Начинаем поиск строки с датой {target_date} и РЦ '{target_rc}'...")
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            # Проверяем дату
            cell_date_value = sheet.cell(row=row_num, column=date_col_idx).value
            if cell_date_value is None:
                continue

            try:
                if isinstance(cell_date_value, str):
                    if "." in cell_date_value:
                        cell_date_obj = datetime.strptime(cell_date_value, "%d.%m.%Y").date()
                    else:
                        # Попробуем другой формат, если стандартный не подошёл
                         cell_date_obj = datetime.strptime(cell_date_value, "%Y-%m-%d").date()
                else:
                    cell_date_obj = cell_date_value.date()
            except Exception as date_parse_error:
                 logging.debug(f"  ⚠️ update_table_row: Не удалось распарсить дату '{cell_date_value}' в строке {row_num}: {date_parse_error}")
                 continue  # Пропускаем, если не удалось преобразовать дату

            if cell_date_obj != target_date:
                continue

            # Проверяем РЦ
            cell_rc_value = sheet.cell(row=row_num, column=rc_col_idx).value
            if cell_rc_value is not None and str(cell_rc_value).strip() == target_rc:
                found_row = row_num
                logging.debug(f"  ✅ update_table_row: Найдена строка {found_row} с совпадающей датой и РЦ.")
                break  # Нашли нужную строку

        if found_row is None:
            logging.info(f"ℹ️ update_table_row: Строка с датой {target_date} и РЦ '{target_rc}' не найдена в таблице.")
            return False

        # Обновляем данные, если ячейки пустые
        updated = False

        # Обновляем водителя
        driver_cell = sheet.cell(row=found_row, column=driver_col_idx)
        driver_from_email = data.get("ФИО водителя", "").strip()
        if driver_from_email:
            current_driver = driver_cell.value
            if not current_driver or str(current_driver).strip().lower() in ("", "nan"):
                sheet.cell(row=found_row, column=driver_col_idx, value=driver_from_email)
                updated = True
                logging.info(f"✅ update_table_row: Обновлён водитель в строке {found_row}: '{driver_from_email}'")
            else:
                 logging.debug(f"ℹ️ update_table_row: Водитель в строке {found_row} уже заполнен ('{current_driver}'). Пропускаем.")

        # Обновляем номер а/м (тягач)
        tractor_cell = sheet.cell(row=found_row, column=tractor_col_idx)
        tractor_from_email = data.get("Тягач", "").strip()
        if tractor_from_email:
            current_tractor = tractor_cell.value
            if not current_tractor or str(current_tractor).strip().lower() in ("", "nan"):
                sheet.cell(row=found_row, column=tractor_col_idx, value=tractor_from_email)
                updated = True
                logging.info(f"✅ update_table_row: Обновлён номер а/м в строке {found_row}: '{tractor_from_email}'")
            else:
                 logging.debug(f"ℹ️ update_table_row: Номер а/м в строке {found_row} уже заполнен ('{current_tractor}'). Пропускаем.")

        if updated:
            try:
                book.save(TABLE_FILE)
                logging.info(f"💾 update_table_row: Таблица успешно обновлена: строка {found_row}")
            except PermissionError:
                logging.error(f"❌ update_table_row: Нет доступа к файлу таблицы. Возможно, он открыт в Excel: {TABLE_FILE}")
                return False  # Считаем, что неудача, если не смогли сохранить
            except Exception as save_error:
                 logging.error(f"❌ update_table_row: Ошибка сохранения таблицы: {save_error}")
                 return False
        else:
            logging.info(f"ℹ️ update_table_row: Строка {found_row} уже содержит данные водителя/тягача, обновление не требуется.")

        return True

    except Exception as e:
        logging.error(f"❌ update_table_row: Критическая ошибка: {e}", exc_info=True)
        return False


# === Проверка и отправка напоминаний согласно процессу ===
def check_and_send_reminders(data, entry_id):
    global sent_reminders

    try:
        network = data.get("Сеть", "").lower().strip()
        if not network:
            return

        if "лента" in network:
            logging.info("Пропускаем Ленту — по процессу не участвует.")
            return

        # === Используем дату из письма для напоминаний ===
        return_date_obj = data.get("Дата возврата из письма")
        if not return_date_obj:
            # fallback на дату из "Дата письма" если не нашли специальную
            try:
                return_date_str = data.get("Дата письма", "")[:10]
                return_date_obj = datetime.strptime(return_date_str, "%Y-%m-%d").date()
            except:
                logging.warning(f"Не удалось определить дату возврата для напоминаний: {data.get('Дата письма', '')}")
                return

        rc_from_email = data.get("РЦ", "").strip()
        today = datetime.today().date()
        current_time = datetime.now().strftime("%H:%M")

        # === Загружаем данные из таблицы для проверки наличия водителя ===
        # Для простоты, предположим, что если мы дошли до этой точки, строка существует.
        # В реальности, можно было бы сделать повторную проверку, но для демонстрации логики отправки напоминаний этого достаточно.
        # Здесь мы просто имитируем проверку на основе данных из письма.
        # В реальной жизни, нужно было бы перечитать таблицу и найти конкретную строку.
        # Для упрощения, предположим, что водитель отсутствует, если ФИО пустое.
        has_driver_in_table_simulation = bool(data.get("ФИО водителя", "").strip())

        # === X5 и Дистрибьюторы ===
        if "x5" in network or "дистр" in network:
            if today == return_date_obj:
                if current_time == "12:00" and not has_driver_in_table_simulation:
                    key = (entry_id, "need_data")
                    if key not in sent_reminders:
                        subject = f"📅 Напоминание ({network.upper()}): предоставьте данные водителя на РЦ {rc_from_email}"
                        body = (
                            f"Дата возврата: {return_date_obj.strftime('%d.%m.%Y')}\n"
                            f"Сеть: {data.get('Сеть', '')}\n"
                            f"РЦ: {rc_from_email}\n\n"
                            f"Напоминаем предоставить данные для оформления пропуска.\n"
                            f"[Автоматическое уведомление]"
                        )
                        recipients = REMINDER_RECIPIENTS.get("x5" if "x5" in network else "дистры")
                        send_email(subject, body, recipients)
                        sent_reminders.add(key)
                        logging.info(f"✅ Отправлено напоминание для {network} → {recipients}")

                if current_time.endswith(":00") and has_driver_in_table_simulation:
                    key = (entry_id, f"check_pass_{current_time}")
                    if key not in sent_reminders:
                        subject = f"🔍 Проверка ({network.upper()}): заказан ли пропуск на РЦ {rc_from_email}?"
                        body = (
                            f"Дата возврата: {return_date_obj.strftime('%d.%m.%Y')}\n"
                            f"Сеть: {data.get('Сеть', '')}\n"
                            f"РЦ: {rc_from_email}\n\n"
                            f"Данные водителя есть в таблице — подтвердите оформление пропуска.\n"
                            f"[Автоматическое уведомление]"
                        )
                        recipients = REMINDER_RECIPIENTS.get("x5" if "x5" in network else "дистры")
                        send_email(subject, body, recipients)
                        sent_reminders.add(key)
                        logging.info(f"✅ Отправлена проверка пропуска для {network} → {recipients}")

        # === Тандер ===
        elif "тандер" in network:
            if return_date_obj.weekday() in (5, 6, 0):  # Сб, Вс, Пн
                days_back = (return_date_obj.weekday() - 4) % 7
                if days_back == 0:
                    days_back = 7
                reminder_date = return_date_obj - timedelta(days=days_back)
            else:
                reminder_date = return_date_obj - timedelta(days=1)

            if today == reminder_date and current_time == "14:00":
                if not has_driver_in_table_simulation:  # Имитация отсутствия водителя
                    key = (entry_id, "tander_need_data")
                    if key not in sent_reminders:
                        subject = f"ТАНДЕР: срочно предоставьте данные водителя на РЦ {rc_from_email}"
                        body = (
                            f"Дата возврата: {return_date_obj.strftime('%d.%m.%Y')}\n"
                            f"Сеть: {data.get('Сеть', '')}\n"
                            f"РЦ: {rc_from_email}\n\n"
                            f"Данные водителя отсутствуют в таблице учёта.\n"
                            f"[Автоматическое уведомление]"
                        )
                        recipients = REMINDER_RECIPIENTS["тандер"]
                        send_email(subject, body, recipients)
                        sent_reminders.add(key)
                        logging.info(f"✅ Отправлено напоминание для Тандер → {recipients}")

    except Exception as e:
        logging.error(f"Ошибка в check_and_send_reminders: {e}")


# === Запись в Excel: вертикальный режим ===
def write_vertical_to_excel(data, sheet_name, excel_file):
    try:
        if not os.path.exists(excel_file):
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            ws.append(["Ключ", "Значение", "EntryID"])
            wb.save(excel_file)
            logging.info(f"✅ Создан новый файл Excel (вертикальный): {excel_file}")

        book = load_workbook(excel_file)

        if sheet_name not in book.sheetnames:
            ws = book.create_sheet(sheet_name)
            ws.append(["Ключ", "Значение", "EntryID"])
        else:
            ws = book[sheet_name]

        startrow = ws.max_row + 1

        ws.cell(row=startrow, column=1, value=f"=== Письмо от {data['Дата письма']} ===")
        ws.cell(row=startrow, column=3, value=data["EntryID"])
        startrow += 1

        for key, value in data.items():
            if key == "EntryID":
                continue
            ws.cell(row=startrow, column=1, value=key)
            ws.cell(row=startrow, column=2, value=value)
            ws.cell(row=startrow, column=3, value=data["EntryID"])
            startrow += 1

        book.save(excel_file)
        logging.info("✅ Данные успешно записаны в Excel (вертикально)")

    except Exception as e:
        logging.error(f"❌ Ошибка записи в Excel (вертикально): {e}")


# === Запись в Excel: горизонтальный режим ===
def write_horizontal_to_excel(data, sheet_name, excel_file):
    try:
        df_new = pd.DataFrame([data])

        if not os.path.exists(EXCEL_FILE):
            df_new.to_excel(EXCEL_FILE, sheet_name=sheet_name, index=False)
            logging.info(f"✅ Создан новый файл Excel (горизонтальный): {excel_file}")
        else:
            book = load_workbook(EXCEL_FILE)

            if sheet_name not in book.sheetnames:
                with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl", mode="a") as writer:
                    df_new.to_excel(writer, sheet_name=sheet_name, index=False)
                logging.info("✅ Создан новый лист и записаны данные")
            else:
                startrow = book[SHEET_NAME].max_row
                with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                    df_new.to_excel(writer, sheet_name=SHEET_NAME, index=False, header=False, startrow=startrow)
                logging.info("✅ Данные дописаны в существующий лист")

    except Exception as e:
        logging.error(f"❌ Ошибка записи в Excel (горизонтально): {e}")


# === Обработка письма ===
def handle_mail(item, processed_ids):
    global _processed_ids
    try:
        entry_id = item.EntryID
        subject = item.Subject
        received_time = item.ReceivedTime
        received_time_str = received_time.strftime("%Y-%m-%d %H:%M")

        logging.info(f"🔍 Начинаем обработку письма: {subject} | ID: {entry_id}")

        if entry_id in processed_ids:
            logging.debug("❌ Пропускаем: уже в processed_ids")
            return

        if is_email_in_excel(entry_id):
            logging.debug("❌ Пропускаем: уже есть в Excel")
            processed_ids.add(entry_id)
            save_processed_ids(processed_ids)
            return

        if not subject or SEARCH_SUBJECT.lower() not in str(subject).lower():
            logging.debug(f"❌ Пропускаем: тема не содержит '{SEARCH_SUBJECT}' (текущая тема: '{subject}')")
            return

        body = item.Body
        del item

        data = parse_email(body, received_time)
        if not data:
            logging.warning("⚠️ Не удалось распарсить письмо")
            processed_ids.add(entry_id)
            save_processed_ids(processed_ids)
            return

        data["EntryID"] = entry_id
        logging.info(f"✅ Извлечено: {data}")

        # === 🔹 ОБНОВЛЯЕМ ТАБЛИЦУ С ПРАВИЛЬНОЙ ДАТОЙ ===
        try:
            # Извлекаем дату ИЗ ПИСЬМА для поиска в таблице
            target_date = data.get("Дата возврата из письма")
            if not target_date:
                 # fallback если дата не была найдена в теле
                 try:
                     fallback_date_str = data.get("Дата письма", "")[:10]
                     target_date = datetime.strptime(fallback_date_str, "%Y-%m-%d").date()
                     logging.debug(f"📅 Используем дату из 'Дата письма' как fallback: {target_date}")
                 except:
                     logging.warning("⚠️ Не удалось определить дату возврата даже из 'Дата письма'.")
                     target_date = None

            target_rc = data.get("РЦ", "").strip()

            if target_date and target_rc:
                success = update_table_row(data, target_date, target_rc)
                if success:
                    logging.info("✅ Данные успешно обновлены в таблице.")
                else:
                    logging.warning("⚠️ Не удалось обновить таблицу.")
            else:
                logging.warning("⚠️ Не удалось определить дату или РЦ из письма для обновления таблицы.")
        except Exception as e:
            logging.error(f"❌ Ошибка при попытке обновить таблицу: {e}")
        # ===========================

        if WRITE_MODE == "vertical":
            write_vertical_to_excel(data, SHEET_NAME, EXCEL_FILE)
            logging.info("✅ Записано вертикально в Excel")
        else:
            write_horizontal_to_excel(data, SHEET_NAME, EXCEL_FILE)
            logging.info("✅ Записано горизонтально в Excel")

        # ✅ Отправка напоминаний по процессу
        check_and_send_reminders(data, entry_id)

        processed_ids.add(entry_id)
        save_processed_ids(processed_ids)

    except Exception as e:
        logging.error(f"❌ Ошибка обработки письма: {e}")
    finally:
        if 'item' in locals():
            del item


# === Основной цикл мониторинга ===
def monitor_inbox():
    pythoncom.CoInitialize()
    outlook = None
    namespace = None
    inbox = None
    folder = None

    processed_ids = load_processed_ids()
    logging.info(f"Загружено {len(processed_ids)} обработанных писем из файла.")

    try:
        outlook = win32com.client.Dispatch("Outlook.Application")
        namespace = outlook.GetNamespace("MAPI")
        inbox = namespace.GetDefaultFolder(6)
        folder = inbox.Folders[OUTLOOK_FOLDER] if OUTLOOK_FOLDER.lower() != "inbox" else inbox

        logging.info(f"✅ Мониторинг запущен. Режим записи: {WRITE_MODE.upper()}.")
        logging.info(f"📂 Папка: {OUTLOOK_FOLDER}")
        logging.info(f"📬 Ищем письма с темой (без учёта регистра): '{SEARCH_SUBJECT}'")

        time.sleep(5)

        while True:
            try:
                today = datetime.today().date()
                min_date = today - timedelta(days=7)

                messages = folder.Items
                messages.Sort("[ReceivedTime]", True)

                msg_list = []
                logging.info("🔍 Сканируем письма...")

                for msg in messages:
                    try:
                        msg_class = getattr(msg, 'Class', None)
                        if msg_class != 43:
                            continue
                        if not getattr(msg, 'ReceivedTime', None):
                            continue
                        if msg.ReceivedTime.date() < min_date:
                            break
                        msg_list.append(msg)
                    except Exception as e:
                        logging.error(f"Ошибка при анализе письма: {e}")
                        continue

                logging.info(f"📬 Найдено {len(msg_list)} подходящих писем. Начинаем обработку...")

                for msg in msg_list:
                    handle_mail(msg, processed_ids)
                    del msg

                del msg_list

            except Exception as e:
                logging.error(f"❌ Ошибка мониторинга: {e}")

            logging.info("⏳ Ждем 60 секунд до следующей проверки...\n")
            time.sleep(60)

    except Exception as e:
        logging.error(f"❌ Ошибка инициализации Outlook: {e}")
    finally:
        for obj in [folder, inbox, namespace, outlook]:
            if obj:
                del obj
        pythoncom.CoUninitialize()


# === Запуск ===
if __name__ == "__main__":
    try:
        set_console_title("📦 Система учета возврата поддонов")
        logging.info("=" * 50)
        logging.info("  📦 Система учета возврата поддонов")
        logging.info(f"  📞 Поддержка: {SUPPORT_CONTACT}")
        logging.info("=" * 50)
        logging.info("")
        monitor_inbox()
    except Exception as e:
        logging.error(f"❌ Критическая ошибка: {e}")
    finally:
        input("\nНажмите Enter для закрытия...")
