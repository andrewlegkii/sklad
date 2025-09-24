import win32com.client
import re
import pandas as pd
from openpyxl import load_workbook, Workbook
import os
from datetime import datetime, timedelta
import time
import pythoncom
import sys
import logging

# === Определяем базовую папку (где лежит .exe или .py) ===
def get_base_path():
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    else:
        return os.path.dirname(os.path.abspath(__file__))

BASE_PATH = get_base_path()

# === Настройка логирования ===
log_file = os.path.join(BASE_PATH, "script.log")
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(log_file, encoding='utf-8'),
        logging.StreamHandler(sys.stdout)
    ]
)

# === Настройки ===
OUTLOOK_FOLDER = "Inbox"
SEARCH_SUBJECT = "Возврат поддонов из сетей"
EXCEL_FILE = os.path.join(BASE_PATH, "возврат_поддонов.xlsx")
PROCESSED_IDS_FILE = os.path.join(BASE_PATH, "processed_ids.txt")
SHEET_NAME = "Данные"

WRITE_MODE = "horizontal"

# === Получатели напоминаний (можно вынести в конфиг позже) ===
# REMINDER_RECIPIENTS = {
#    "x5": ["dma@line7.ru", "slon07@line7.ru", "rudcekb@nestlesoft.net"],
#    "тандер": ["rudcekb@nestlesoft.net"],
#    "дистры": ["rudcekb@nestlesoft.net"]
# }

REMINDER_RECIPIENTS = {
    "x5": ["skoppss@yandex.ru"],
    "тандер": ["skoppss@yandex.ru"],
    "дистры": ["skoppss@yandex.ru"]
}

# === Служебные переменные для отслеживания отправленных напоминаний ===
sent_reminders = set()  # Хранит (entry_id, тип_напоминания)


# === Загрузка уже обработанных ID из файла ===
def load_processed_ids():
    if not os.path.exists(PROCESSED_IDS_FILE):
        return set()
    try:
        with open(PROCESSED_IDS_FILE, "r", encoding="utf-8") as f:
            return set(line.strip() for line in f if line.strip())
    except Exception as e:
        logging.error(f"Ошибка загрузки processed_ids: {e}")
        return set()


# === Сохранение обработанных ID в файл ===
def save_processed_ids(ids):
    try:
        with open(PROCESSED_IDS_FILE, "w", encoding="utf-8") as f:
            for item_id in ids:
                f.write(item_id + "\n")
    except Exception as e:
        logging.error(f"Ошибка сохранения processed_ids: {e}")


# === Проверка, есть ли письмо уже в Excel (по EntryID) ===
def is_email_in_excel(entry_id):
    if not os.path.exists(EXCEL_FILE):
        return False
    try:
        df = pd.read_excel(EXCEL_FILE, sheet_name=SHEET_NAME)
        if "EntryID" in df.columns and entry_id in df["EntryID"].values:
            return True
    except Exception as e:
        logging.debug(f"Ошибка проверки Excel: {e}")
    return False


# === Парсинг письма ===
def parse_email(body, received_time):
    try:
        lines = body.splitlines()
        data = {
            "Дата письма": received_time.strftime("%Y-%m-%d %H:%M"),
            "Сеть": "",
            "РЦ": "",
            "Тягач": "",
            "Прицеп": "",
            "ФИО водителя": "",
            "Паспорт": "",
            "Номер ВУ": "",
            "Телефон": "",
            "ИНН": "",
            "Доп. информация": "",
            "EntryID": ""
        }

        for line in lines:
            line = line.strip()
            if not line:
                continue

            if line.startswith("Сеть"):
                parts = [p.strip() for p in line.split("|")]
                if len(parts) >= 2:
                    data["Сеть"] = parts[1]
                if len(parts) >= 3:
                    data["РЦ"] = parts[2].replace("РЦ", "").strip()
            elif line.startswith("Тягач"):
                data["Тягач"] = line.split(":", 1)[1].strip() if ":" in line else ""
            elif line.startswith("Прицеп"):
                data["Прицеп"] = line.split(":", 1)[1].strip() if ":" in line else ""
            elif line.startswith("Ф.И.О. водителя"):
                data["ФИО водителя"] = line.split(":", 1)[1].strip() if ":" in line else ""
            elif line.startswith("Паспорт"):
                data["Паспорт"] = line.split(":", 1)[1].strip() if ":" in line else ""
            elif line.startswith("Номер ВУ"):
                data["Номер ВУ"] = line.split(":", 1)[1].strip() if ":" in line else ""
            elif line.startswith("Телефон"):
                data["Телефон"] = line.split(":", 1)[1].strip() if ":" in line else ""
            elif line.startswith("ИНН"):
                data["ИНН"] = line.split(":", 1)[1].strip() if ":" in line else ""
            elif line.startswith("Дополнительная информация"):
                data["Доп. информация"] = line.split(":", 1)[1].strip() if ":" in line else ""

        return data

    except Exception as e:
        logging.error(f"Ошибка парсинга письма: {e}")
        return None


# === Отправка уведомления ===
def send_email(subject, body, to, cc=None):
    try:
        outlook_app = win32com.client.Dispatch("Outlook.Application")
        mail = outlook_app.CreateItem(0)
        mail.Subject = subject
        if isinstance(to, list):
            mail.To = ";".join(to)
        else:
            mail.To = to
        if cc:
            mail.CC = cc
        mail.Body = body
        mail.Send()
        logging.info(f"Отправлено уведомление: {subject} -> {to}")
        del mail
        del outlook_app
    except Exception as e:
        logging.error(f"Ошибка отправки email: {e}")


# === Проверка и отправка напоминаний согласно процессу ===
def check_and_send_reminders(data, entry_id):
    global sent_reminders

    try:
        # Определяем сеть (в нижнем регистре для сравнения)
        network = data.get("Сеть", "").lower().strip()
        if not network:
            return

        # Пропускаем "Лента" — по процессу не участвует
        if "лента" in network:
            logging.info("Пропускаем напоминания для Ленты — по процессу не участвует.")
            return

        # Пытаемся получить дату возврата из письма
        return_date_str = data.get("Дата письма", "")[:10]  # Берём только YYYY-MM-DD
        try:
            return_date = datetime.strptime(return_date_str, "%Y-%m-%d").date()
        except:
            logging.warning(f"Не удалось определить дату возврата из письма: {return_date_str}")
            return

        today = datetime.today().date()
        current_time = datetime.now().strftime("%H:%M")

        # === Логика для X5 и Дистрибьюторов ===
        if "x5" in network or "дистр" in network:
            # Напоминание в 12:00 в день возврата
            if today == return_date and current_time == "12:00":
                reminder_key = (entry_id, "due_day_1200")
                if reminder_key not in sent_reminders:
                    subject = f"📅 Напоминание ({network.upper()}): предоставить данные для пропуска на РЦ"
                    body = (
                        f"Дата возврата: {return_date_str}\n"
                        f"Сеть: {data.get('Сеть', '')}\n"
                        f"РЦ: {data.get('РЦ', '')}\n\n"
                        f"Напоминаем предоставить данные для оформления пропуска.\n"
                        f"[Автоматическое уведомление]"
                    )
                    recipients = REMINDER_RECIPIENTS.get("x5" if "x5" in network else "дистры")
                    send_email(subject, body, recipients)
                    sent_reminders.add(reminder_key)
                    logging.info(f"✅ Отправлено напоминание для {network} в день возврата → {recipients}")

            # Проверка отправки пропуска — раз в час в день возврата
            if today == return_date and int(current_time[3:]) == 0:  # Каждый час:XX, где XX=00
                reminder_key = (entry_id, f"hourly_check_{current_time[:2]}")
                if reminder_key not in sent_reminders:
                    subject = f"🔍 Проверка ({network.upper()}): отправлен ли пропуск на РЦ?"
                    body = (
                        f"Дата возврата: {return_date_str}\n"
                        f"Сеть: {data.get('Сеть', '')}\n"
                        f"РЦ: {data.get('РЦ', '')}\n\n"
                        f"Пожалуйста, подтвердите, что пропуск на РЦ оформлен и отправлен.\n"
                        f"[Автоматическое уведомление]"
                    )
                    recipients = REMINDER_RECIPIENTS.get("x5" if "x5" in network else "дистры")
                    send_email(subject, body, recipients)
                    sent_reminders.add(reminder_key)
                    logging.info(f"✅ Отправлена проверка пропуска для {network} → {recipients}")

        # === Логика для Тандер ===
        elif "тандер" in network:
            # Определяем "канун" — день перед возвратом
            due_eve = return_date - timedelta(days=1)

            # Если возврат в понедельник — напоминать в пятницу
            if return_date.weekday() == 0:  # Понедельник
                due_eve = return_date - timedelta(days=3)

            # Напоминание в 14:00 в канун дня возврата
            if today == due_eve and current_time == "14:00":
                reminder_key = (entry_id, "due_eve_1400")
                if reminder_key not in sent_reminders:
                    subject = "🚛 Напоминание (ТАНДЕР): предоставить данные для пропуска на РЦ"
                    body = (
                        f"Дата возврата: {return_date_str}\n"
                        f"Сеть: {data.get('Сеть', '')}\n"
                        f"РЦ: {data.get('РЦ', '')}\n\n"
                        f"Напоминаем предоставить данные для оформления пропуска НАКАНУНЕ возврата.\n"
                        f"[Автоматическое уведомление]"
                    )
                    recipients = REMINDER_RECIPIENTS.get("тандер")
                    send_email(subject, body, recipients)
                    sent_reminders.add(reminder_key)
                    logging.info(f"✅ Отправлено напоминание для Тандер (накануне) → {recipients}")

    except Exception as e:
        logging.error(f"Ошибка при отправке напоминания: {e}")


# === Запись в Excel: вертикальный режим ===
def write_vertical_to_excel(data, sheet_name, excel_file):
    try:
        if not os.path.exists(excel_file):
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            ws.append(["Ключ", "Значение", "EntryID"])
            wb.save(excel_file)
            logging.info(f"✅ Создан новый файл Excel (вертикальный): {excel_file}")

        book = load_workbook(excel_file)

        if sheet_name not in book.sheetnames:
            ws = book.create_sheet(sheet_name)
            ws.append(["Ключ", "Значение", "EntryID"])
        else:
            ws = book[sheet_name]

        startrow = ws.max_row + 1

        ws.cell(row=startrow, column=1, value=f"=== Письмо от {data['Дата письма']} ===")
        ws.cell(row=startrow, column=3, value=data["EntryID"])
        startrow += 1

        for key, value in data.items():
            if key == "EntryID":
                continue
            ws.cell(row=startrow, column=1, value=key)
            ws.cell(row=startrow, column=2, value=value)
            ws.cell(row=startrow, column=3, value=data["EntryID"])
            startrow += 1

        book.save(excel_file)
        logging.info("✅ Данные успешно записаны в Excel (вертикально)")

    except Exception as e:
        logging.error(f"❌ Ошибка записи в Excel (вертикально): {e}")


# === Запись в Excel: горизонтальный режим ===
def write_horizontal_to_excel(data, sheet_name, excel_file):
    try:
        df_new = pd.DataFrame([data])

        if not os.path.exists(EXCEL_FILE):
            df_new.to_excel(excel_file, sheet_name=sheet_name, index=False)
            logging.info(f"✅ Создан новый файл Excel (горизонтальный): {excel_file}")
        else:
            book = load_workbook(excel_file)

            if sheet_name not in book.sheetnames:
                with pd.ExcelWriter(excel_file, engine="openpyxl", mode="a") as writer:
                    df_new.to_excel(writer, sheet_name=sheet_name, index=False)
                logging.info("✅ Создан новый лист и записаны данные")
            else:
                startrow = book[sheet_name].max_row
                with pd.ExcelWriter(excel_file, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                    df_new.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=startrow)
                logging.info("✅ Данные дописаны в существующий лист")

    except Exception as e:
        logging.error(f"❌ Ошибка записи в Excel (горизонтально): {e}")


# === Обработка письма ===
def handle_mail(item, processed_ids):
    try:
        entry_id = item.EntryID
        subject = item.Subject
        received_time = item.ReceivedTime
        received_time_str = received_time.strftime("%Y-%m-%d %H:%M")

        logging.info(f"🔍 Начинаем обработку письма: {subject} | ID: {entry_id}")

        if entry_id in processed_ids:
            logging.debug("❌ Пропускаем: уже в processed_ids")
            return

        if is_email_in_excel(entry_id):
            logging.debug("❌ Пропускаем: уже есть в Excel")
            processed_ids.add(entry_id)
            save_processed_ids(processed_ids)
            return

        if not subject or SEARCH_SUBJECT.lower() not in str(subject).lower():
            logging.debug(f"❌ Пропускаем: тема не содержит '{SEARCH_SUBJECT}' (текущая тема: '{subject}')")
            return

        body = item.Body
        del item

        data = parse_email(body, received_time)
        if not data:
            logging.warning("⚠️ Не удалось распарсить письмо")
            processed_ids.add(entry_id)
            save_processed_ids(processed_ids)
            return

        data["EntryID"] = entry_id
        logging.info(f"✅ Извлечено: {data}")

        if WRITE_MODE == "vertical":
            write_vertical_to_excel(data, SHEET_NAME, EXCEL_FILE)
            logging.info("✅ Записано вертикально в Excel")
        else:
            write_horizontal_to_excel(data, SHEET_NAME, EXCEL_FILE)
            logging.info("✅ Записано горизонтально в Excel")

        # ✅ Проверяем и отправляем напоминания согласно процессу
        check_and_send_reminders(data, entry_id)

        processed_ids.add(entry_id)
        save_processed_ids(processed_ids)

    except Exception as e:
        logging.error(f"❌ Ошибка обработки письма: {e}")
    finally:
        if 'item' in locals():
            del item


# === Основной цикл мониторинга ===
def monitor_inbox():
    pythoncom.CoInitialize()
    outlook = None
    namespace = None
    inbox = None
    folder = None

    processed_ids = load_processed_ids()
    logging.info(f"Загружено {len(processed_ids)} обработанных писем из файла.")

    try:
        outlook = win32com.client.Dispatch("Outlook.Application")
        namespace = outlook.GetNamespace("MAPI")
        inbox = namespace.GetDefaultFolder(6)
        folder = inbox.Folders[OUTLOOK_FOLDER] if OUTLOOK_FOLDER.lower() != "inbox" else inbox

        logging.info(f"✅ Мониторинг запущен. Режим записи: {WRITE_MODE.upper()}.")
        logging.info(f"📂 Папка: {OUTLOOK_FOLDER}")
        logging.info(f"📬 Ищем письма с темой (без учёта регистра): '{SEARCH_SUBJECT}'")

        time.sleep(5)

        while True:
            try:
                today = datetime.today().date()
                min_date = today - timedelta(days=7)  # Смотрим письма за последнюю неделю

                messages = folder.Items
                messages.Sort("[ReceivedTime]", True)

                msg_list = []
                logging.info("🔍 Сканируем письма...")

                for msg in messages:
                    try:
                        msg_class = getattr(msg, 'Class', None)
                        subject = getattr(msg, 'Subject', 'Без темы')
                        received_time = getattr(msg, 'ReceivedTime', None)

                        if msg_class != 43:
                            continue

                        if not received_time:
                            continue

                        if received_time.date() < min_date:
                            break

                        msg_list.append(msg)

                    except Exception as e:
                        logging.error(f"Ошибка при анализе письма: {e}")
                        continue

                logging.info(f"📬 Найдено {len(msg_list)} подходящих писем. Начинаем обработку...")

                for msg in msg_list:
                    handle_mail(msg, processed_ids)
                    del msg

                del msg_list

            except Exception as e:
                logging.error(f"❌ Ошибка мониторинга: {e}")

            logging.info("⏳ Ждем 60 секунд до следующей проверки...\n")
            time.sleep(60)

    except Exception as e:
        logging.error(f"❌ Ошибка инициализации Outlook: {e}")
    finally:
        for obj in [folder, inbox, namespace, outlook]:
            if obj:
                del obj
        pythoncom.CoUninitialize()


# === Запуск ===
if __name__ == "__main__":
    try:
        monitor_inbox()
    except Exception as e:
        logging.error(f"❌ Критическая ошибка: {e}")
    finally:
        input("\nНажмите Enter для закрытия...")
