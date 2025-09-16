import win32com.client
import re
import pandas as pd
from openpyxl import load_workbook
import os
from datetime import datetime, timedelta
import time
import pythoncom

# === Настройки ===
OUTLOOK_FOLDER = "Inbox"
SEARCH_SUBJECT = "Возврат поддонов из сетей"
EXCEL_FILE = r"C:\Users\legki\Desktop\test.xlsx"
SHEET_NAME = "Данные"

EMAIL_TO = "skoppss@yandex.ru"
EMAIL_CC = "legkiy.a@inbox.eu"

processed_ids = set()


# === Парсинг письма ===
def parse_email(body, received_time):
    try:
        lines = body.splitlines()
        data = {
            "Дата письма": received_time.strftime("%Y-%m-%d %H:%M"),
            "Сеть": "",
            "РЦ": "",
            "Тягач": "",
            "Прицеп": "",
            "ФИО водителя": "",
            "Паспорт": "",
            "Номер ВУ": "",
            "Телефон": "",
            "ИНН": "",
            "Доп. информация": ""
        }

        for line in lines:
            line = line.strip()
            if not line:
                continue

            if line.startswith("Сеть"):
                parts = [p.strip() for p in line.split("|")]
                if len(parts) >= 2:
                    data["Сеть"] = parts[1]
                if len(parts) >= 3:
                    data["РЦ"] = parts[2].replace("РЦ", "").strip()
            elif line.startswith("Тягач"):
                data["Тягач"] = line.split(":", 1)[1].strip() if ":" in line else ""
            elif line.startswith("Прицеп"):
                data["Прицеп"] = line.split(":", 1)[1].strip() if ":" in line else ""
            elif line.startswith("Ф.И.О. водителя"):
                data["ФИО водителя"] = line.split(":", 1)[1].strip() if ":" in line else ""
            elif line.startswith("Паспорт"):
                data["Паспорт"] = line.split(":", 1)[1].strip() if ":" in line else ""
            elif line.startswith("Номер ВУ"):
                data["Номер ВУ"] = line.split(":", 1)[1].strip() if ":" in line else ""
            elif line.startswith("Телефон"):
                data["Телефон"] = line.split(":", 1)[1].strip() if ":" in line else ""
            elif line.startswith("ИНН"):
                data["ИНН"] = line.split(":", 1)[1].strip() if ":" in line else ""
            elif line.startswith("Дополнительная информация"):
                data["Доп. информация"] = line.split(":", 1)[1].strip() if ":" in line else ""

        return data

    except Exception as e:
        print("Ошибка парсинга письма:", e)
        return None


# === Отправка уведомления ===
def send_email(subject, body, to, cc=None):
    outlook_app = win32com.client.Dispatch("Outlook.Application")
    mail = outlook_app.CreateItem(0)
    mail.Subject = subject
    mail.To = to
    if cc:
        mail.CC = cc
    mail.Body = body
    mail.Send()
    print(f"Отправлено уведомление: {subject} -> {to}")


# === Обработка письма ===
def handle_mail(item):
    try:
        if item.EntryID in processed_ids:
            return
        if not item.Subject or SEARCH_SUBJECT not in str(item.Subject):
            return

        data = parse_email(item.Body, item.ReceivedTime)
        if not data:
            processed_ids.add(item.EntryID)
            return

        print("Извлечено:", data)

        # === Запись в Excel ===
        df_new = pd.DataFrame([data])
        if not os.path.exists(EXCEL_FILE):
            df_new.to_excel(EXCEL_FILE, sheet_name=SHEET_NAME, index=False)
        else:
            book = load_workbook(EXCEL_FILE)
            if SHEET_NAME not in book.sheetnames:
                with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl", mode="a") as writer:
                    df_new.to_excel(writer, sheet_name=SHEET_NAME, index=False)
            else:
                startrow = book[SHEET_NAME].max_row
                with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                    df_new.to_excel(writer, sheet_name=SHEET_NAME, index=False,
                                    header=False, startrow=startrow)

        # === Уведомление для Тандер ===
        if data.get("Сеть", "").lower() == "тандер" and "Дата письма" in data:
            try:
                return_date = datetime.strptime(data["Дата письма"], "%Y-%m-%d %H:%M")
                today_date = datetime.today()
                notify_days = [today_date]

                if today_date.weekday() == 4:  # пятница
                    notify_days = [today_date + timedelta(days=i) for i in [1, 2, 3]]

                for notify_day in notify_days:
                    subject = f"Напоминание: заказать пропуск на РЦ для Тандер"
                    body_msg = (
                        f"Дата возврата: {data['Дата письма']}\n"
                        f"Сеть: {data['Сеть']}\n"
                        f"РЦ: {data['РЦ']}\n"
                        f"Тягач: {data.get('Тягач','')}\n"
                        f"Прицеп: {data.get('Прицеп','')}\n"
                        f"Не забудьте заказать пропуск на РЦ!"
                    )
                    send_email(subject, body_msg, EMAIL_TO, EMAIL_CC)
            except Exception as e:
                print("Ошибка при отправке уведомления:", e)

        processed_ids.add(item.EntryID)
    except Exception as e:
        print("Ошибка обработки письма:", e)


# === Основной цикл мониторинга ===
def monitor_inbox():
    pythoncom.CoInitialize()
    try:
        outlook = win32com.client.GetActiveObject("Outlook.Application")
    except Exception:
        print("Не удалось подключиться к открытому Outlook. Запустите Outlook и войдите в аккаунт.")
        return

    namespace = outlook.GetNamespace("MAPI")
    inbox = namespace.GetDefaultFolder(6)
    folder = inbox.Folders[OUTLOOK_FOLDER] if OUTLOOK_FOLDER.lower() != "inbox" else inbox

    print("Мониторинг запущен. Будут обрабатываться только письма за сегодня.")

    while True:
        try:
            today = datetime.today().date()
            messages = folder.Items
            messages.Sort("[ReceivedTime]", True)

            for msg in messages:
                if msg.ReceivedTime.date() < today:
                    break
                handle_mail(msg)

        except Exception as e:
            print("Ошибка мониторинга:", e)

        print("Ждем 300 секунд до следующей проверки...\n")
        time.sleep(60)


if __name__ == "__main__":
    monitor_inbox()
