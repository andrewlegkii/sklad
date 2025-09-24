import win32com.client
import re
import pandas as pd
from openpyxl import load_workbook, Workbook
import os
from datetime import datetime, timedelta
import time
import pythoncom
import sys
import logging

# === Определяем базовую папку (где лежит .exe или .py) ===
def get_base_path():
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    else:
        return os.path.dirname(os.path.abspath(__file__))

BASE_PATH = get_base_path()

# === ⚠️ УКАЖИ ПУТЬ К ТАБЛИЦЕ С ДАННЫМИ ===
TABLE_FILE = r"C:\Путь\К\Екатеринбург - учет оборота поддонов.xlsx"  # ← ЗАМЕНИ НА СВОЙ ПУТЬ

# === Настройка логирования ===
log_file = os.path.join(BASE_PATH, "script.log")
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(log_file, encoding='utf-8'),
        logging.StreamHandler(sys.stdout)
    ]
)

# === Настройки ===
OUTLOOK_FOLDER = "Inbox"
SEARCH_SUBJECT = "Возврат поддонов из сетей"
EXCEL_FILE = os.path.join(BASE_PATH, "возврат_поддонов.xlsx")
PROCESSED_IDS_FILE = os.path.join(BASE_PATH, "processed_ids.txt")
SHEET_NAME = "Данные"

WRITE_MODE = "horizontal"

# === Получатели напоминаний ===
REMINDER_RECIPIENTS = {
    "x5": ["dma@line7.ru", "slon07@line7.ru", "rudcekb@nestlesoft.net"],
    "тандер": ["rudcekb@nestlesoft.net"],
    "дистры": ["rudcekb@nestlesoft.net"]
}

# === Кэш таблицы (чтобы не читать каждый раз) ===
_table_cache = None
_table_cache_time = None

# === Отправленные напоминания ===
sent_reminders = set()


# === Загрузка данных из таблицы ===
def load_table_data():
    global _table_cache, _table_cache_time
    now = time.time()
    if _table_cache is not None and _table_cache_time > now - 300:  # обновляем раз в 5 мин
        return _table_cache

    try:
        if not os.path.exists(TABLE_FILE):
            logging.warning(f"Таблица не найдена: {TABLE_FILE}")
            _table_cache = []
            _table_cache_time = now
            return _table_cache

        df = pd.read_excel(TABLE_FILE, header=0)
        df.columns = df.columns.str.strip()
        records = []

        for _, row in df.iterrows():
            try:
                date_val = row.get("дата")
                if pd.isna(date_val):
                    continue

                # Преобразуем дату
                if isinstance(date_val, str):
                    if "." in date_val:
                        date_obj = datetime.strptime(date_val, "%d.%m.%Y").date()
                    else:
                        continue
                else:
                    date_obj = date_val.date()

                provider = str(row.get("Поставщик", "")).strip()
                driver = str(row.get("водитель Фамилия И.О.", "")).strip()
                rc = str(row.get("РЦ", "")).strip()

                # Определяем клиента
                client = None
                provider_lower = provider.lower()
                if "эксиси" in provider_lower or "пт групп" in provider_lower or "эксисси" in provider_lower:
                    client = "дистры"
                elif "пермь" in provider_lower or "х5" in provider_lower:
                    client = "x5"
                elif "тандер" in provider_lower:
                    client = "тандер"
                # Добавь другие поставщики по мере необходимости

                records.append({
                    "date": date_obj,
                    "client": client,
                    "rc": rc,
                    "has_driver": bool(driver and driver.lower() not in ("nan", "", "none")),
                    "provider": provider
                })
            except Exception as e:
                logging.debug(f"Пропущена строка таблицы: {e}")
                continue

        _table_cache = records
        _table_cache_time = now
        logging.info(f"Загружено {len(records)} записей из таблицы")
        return records

    except Exception as e:
        logging.error(f"Ошибка чтения таблицы: {e}")
        return []


# === Загрузка уже обработанных ID из файла ===
def load_processed_ids():
    if not os.path.exists(PROCESSED_IDS_FILE):
        return set()
    try:
        with open(PROCESSED_IDS_FILE, "r", encoding="utf-8") as f:
            return set(line.strip() for line in f if line.strip())
    except Exception as e:
        logging.error(f"Ошибка загрузки processed_ids: {e}")
        return set()


# === Сохранение обработанных ID в файл ===
def save_processed_ids(ids):
    try:
        with open(PROCESSED_IDS_FILE, "w", encoding="utf-8") as f:
            for item_id in ids:
                f.write(item_id + "\n")
    except Exception as e:
        logging.error(f"Ошибка сохранения processed_ids: {e}")


# === Проверка, есть ли письмо уже в Excel (по EntryID) ===
def is_email_in_excel(entry_id):
    if not os.path.exists(EXCEL_FILE):
        return False
    try:
        df = pd.read_excel(EXCEL_FILE, sheet_name=SHEET_NAME)
        if "EntryID" in df.columns and entry_id in df["EntryID"].values:
            return True
    except Exception as e:
        logging.debug(f"Ошибка проверки Excel: {e}")
    return False


# === Парсинг письма ===
def parse_email(body, received_time):
    try:
        lines = body.splitlines()
        data = {
            "Дата письма": received_time.strftime("%Y-%m-%d %H:%M"),
            "Сеть": "",
            "РЦ": "",
            "Тягач": "",
            "Прицеп": "",
            "ФИО водителя": "",
            "Паспорт": "",
            "Номер ВУ": "",
            "Телефон": "",
            "ИНН": "",
            "Доп. информация": "",
            "EntryID": ""
        }

        for line in lines:
            line = line.strip()
            if not line:
                continue

            if line.startswith("Сеть"):
                parts = [p.strip() for p in line.split("|")]
                if len(parts) >= 2:
                    data["Сеть"] = parts[1]
                if len(parts) >= 3:
                    data["РЦ"] = parts[2].replace("РЦ", "").strip()
            elif line.startswith("Тягач"):
                data["Тягач"] = line.split(":", 1)[1].strip() if ":" in line else ""
            elif line.startswith("Прицеп"):
                data["Прицеп"] = line.split(":", 1)[1].strip() if ":" in line else ""
            elif line.startswith("Ф.И.О. водителя"):
                data["ФИО водителя"] = line.split(":", 1)[1].strip() if ":" in line else ""
            elif line.startswith("Паспорт"):
                data["Паспорт"] = line.split(":", 1)[1].strip() if ":" in line else ""
            elif line.startswith("Номер ВУ"):
                data["Номер ВУ"] = line.split(":", 1)[1].strip() if ":" in line else ""
            elif line.startswith("Телефон"):
                data["Телефон"] = line.split(":", 1)[1].strip() if ":" in line else ""
            elif line.startswith("ИНН"):
                data["ИНН"] = line.split(":", 1)[1].strip() if ":" in line else ""
            elif line.startswith("Дополнительная информация"):
                data["Доп. информация"] = line.split(":", 1)[1].strip() if ":" in line else ""

        return data

    except Exception as e:
        logging.error(f"Ошибка парсинга письма: {e}")
        return None


# === Отправка уведомления ===
def send_email(subject, body, to, cc=None):
    try:
        outlook_app = win32com.client.Dispatch("Outlook.Application")
        mail = outlook_app.CreateItem(0)
        mail.Subject = subject
        if isinstance(to, list):
            mail.To = ";".join(to)
        else:
            mail.To = to
        if cc:
            mail.CC = cc
        mail.Body = body
        mail.Send()
        logging.info(f"Отправлено уведомление: {subject} -> {to}")
        del mail
        del outlook_app
    except Exception as e:
        logging.error(f"Ошибка отправки email: {e}")


# === Проверка и отправка напоминаний согласно процессу ===
def check_and_send_reminders(data, entry_id):
    global sent_reminders

    try:
        network = data.get("Сеть", "").lower().strip()
        if not network:
            return

        if "лента" in network:
            logging.info("Пропускаем Ленту — по процессу не участвует.")
            return

        # Дата возврата из письма
        return_date_str = data.get("Дата письма", "")[:10]
        try:
            return_date = datetime.strptime(return_date_str, "%Y-%m-%d").date()
        except:
            logging.warning(f"Не удалось определить дату возврата: {return_date_str}")
            return

        rc_from_email = data.get("РЦ", "").strip()
        today = datetime.today().date()
        current_time = datetime.now().strftime("%H:%M")

        # === Загружаем данные из таблицы ===
        table_records = load_table_data()
        has_driver_in_table = None
        client_from_table = None

        for rec in table_records:
            if rec["date"] == return_date and rec["rc"] == rc_from_email:
                has_driver_in_table = rec["has_driver"]
                client_from_table = rec["client"]
                break

        # Если не нашли в таблице — пропускаем
        if client_from_table is None:
            logging.debug(f"Не найдено совпадение в таблице: дата={return_date}, РЦ={rc_from_email}")
            return

        # Используем клиента из таблицы
        network = client_from_table

        # === X5 и Дистрибьюторы ===
        if network in ("x5", "дистры"):
            if today == return_date:
                # 12:00 — напоминание, если водителя нет
                if current_time == "12:00" and not has_driver_in_table:
                    key = (entry_id, "need_data")
                    if key not in sent_reminders:
                        subject = f"📅 Напоминание ({network.upper()}): предоставьте данные водителя на РЦ {rc_from_email}"
                        body = (
                            f"Дата возврата: {return_date.strftime('%d.%m.%Y')}\n"
                            f"РЦ: {rc_from_email}\n\n"
                            f"Данные водителя отсутствуют в таблице учёта.\n"
                            f"[Автоматическое уведомление]"
                        )
                        recipients = REMINDER_RECIPIENTS[network]
                        send_email(subject, body, recipients)
                        sent_reminders.add(key)
                        logging.info(f"✅ Отправлено напоминание о данных для {network} → {recipients}")

                # Каждый час — проверка пропуска, если водитель есть
                if current_time.endswith(":00") and has_driver_in_table:
                    key = (entry_id, f"check_pass_{current_time}")
                    if key not in sent_reminders:
                        subject = f"🔍 Проверка ({network.upper()}): заказан ли пропуск на РЦ {rc_from_email}?"
                        body = (
                            f"Дата возврата: {return_date.strftime('%d.%m.%Y')}\n"
                            f"РЦ: {rc_from_email}\n\n"
                            f"Данные водителя есть в таблице — подтвердите оформление пропуска.\n"
                            f"[Автоматическое уведомление]"
                        )
                        recipients = REMINDER_RECIPIENTS[network]
                        send_email(subject, body, recipients)
                        sent_reminders.add(key)
                        logging.info(f"✅ Отправлена проверка пропуска для {network} → {recipients}")

        # === Тандер ===
        elif network == "тандер":
            if return_date.weekday() in (5, 6, 0):  # Сб, Вс, Пн
                days_back = (return_date.weekday() - 4) % 7
                if days_back == 0:
                    days_back = 7
                reminder_date = return_date - timedelta(days=days_back)
            else:
                reminder_date = return_date - timedelta(days=1)

            if today == reminder_date and current_time == "14:00":
                if not has_driver_in_table:
                    key = (entry_id, "tander_need_data")
                    if key not in sent_reminders:
                        subject = f"🚛 ТАНДЕР: срочно предоставьте данные водителя на РЦ {rc_from_email}"
                        body = (
                            f"Дата возврата: {return_date.strftime('%d.%m.%Y')}\n"
                            f"РЦ: {rc_from_email}\n\n"
                            f"Данные водителя отсутствуют в таблице учёта.\n"
                            f"[Автоматическое уведомление]"
                        )
                        recipients = REMINDER_RECIPIENTS["тандер"]
                        send_email(subject, body, recipients)
                        sent_reminders.add(key)
                        logging.info(f"✅ Отправлено напоминание для Тандер → {recipients}")

    except Exception as e:
        logging.error(f"Ошибка в check_and_send_reminders: {e}")


# === Запись в Excel: вертикальный режим ===
def write_vertical_to_excel(data, sheet_name, excel_file):
    try:
        if not os.path.exists(excel_file):
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            ws.append(["Ключ", "Значение", "EntryID"])
            wb.save(excel_file)
            logging.info(f"✅ Создан новый файл Excel (вертикальный): {excel_file}")

        book = load_workbook(excel_file)

        if sheet_name not in book.sheetnames:
            ws = book.create_sheet(sheet_name)
            ws.append(["Ключ", "Значение", "EntryID"])
        else:
            ws = book[sheet_name]

        startrow = ws.max_row + 1

        ws.cell(row=startrow, column=1, value=f"=== Письмо от {data['Дата письма']} ===")
        ws.cell(row=startrow, column=3, value=data["EntryID"])
        startrow += 1

        for key, value in data.items():
            if key == "EntryID":
                continue
            ws.cell(row=startrow, column=1, value=key)
            ws.cell(row=startrow, column=2, value=value)
            ws.cell(row=startrow, column=3, value=data["EntryID"])
            startrow += 1

        book.save(excel_file)
        logging.info("✅ Данные успешно записаны в Excel (вертикально)")

    except Exception as e:
        logging.error(f"❌ Ошибка записи в Excel (вертикально): {e}")


# === Запись в Excel: горизонтальный режим ===
def write_horizontal_to_excel(data, sheet_name, excel_file):
    try:
        df_new = pd.DataFrame([data])

        if not os.path.exists(EXCEL_FILE):
            df_new.to_excel(EXCEL_FILE, sheet_name=sheet_name, index=False)
            logging.info(f"✅ Создан новый файл Excel (горизонтальный): {excel_file}")
        else:
            book = load_workbook(EXCEL_FILE)

            if sheet_name not in book.sheetnames:
                with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl", mode="a") as writer:
                    df_new.to_excel(writer, sheet_name=sheet_name, index=False)
                logging.info("✅ Создан новый лист и записаны данные")
            else:
                startrow = book[SHEET_NAME].max_row
                with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                    df_new.to_excel(writer, sheet_name=SHEET_NAME, index=False, header=False, startrow=startrow)
                logging.info("✅ Данные дописаны в существующий лист")

    except Exception as e:
        logging.error(f"❌ Ошибка записи в Excel (горизонтально): {e}")


# === Обработка письма ===
def handle_mail(item, processed_ids):
    try:
        entry_id = item.EntryID
        subject = item.Subject
        received_time = item.ReceivedTime
        received_time_str = received_time.strftime("%Y-%m-%d %H:%M")

        logging.info(f"🔍 Начинаем обработку письма: {subject} | ID: {entry_id}")

        if entry_id in processed_ids:
            logging.debug("❌ Пропускаем: уже в processed_ids")
            return

        if is_email_in_excel(entry_id):
            logging.debug("❌ Пропускаем: уже есть в Excel")
            processed_ids.add(entry_id)
            save_processed_ids(processed_ids)
            return

        if not subject or SEARCH_SUBJECT.lower() not in str(subject).lower():
            logging.debug(f"❌ Пропускаем: тема не содержит '{SEARCH_SUBJECT}' (текущая тема: '{subject}')")
            return

        body = item.Body
        del item

        data = parse_email(body, received_time)
        if not 
            logging.warning("⚠️ Не удалось распарсить письмо")
            processed_ids.add(entry_id)
            save_processed_ids(processed_ids)
            return

        data["EntryID"] = entry_id
        logging.info(f"✅ Извлечено: {data}")

        if WRITE_MODE == "vertical":
            write_vertical_to_excel(data, SHEET_NAME, EXCEL_FILE)
            logging.info("✅ Записано вертикально в Excel")
        else:
            write_horizontal_to_excel(data, SHEET_NAME, EXCEL_FILE)
            logging.info("✅ Записано горизонтально в Excel")

        # ✅ Отправка напоминаний по процессу с учётом таблицы
        check_and_send_reminders(data, entry_id)

        processed_ids.add(entry_id)
        save_processed_ids(processed_ids)

    except Exception as e:
        logging.error(f"❌ Ошибка обработки письма: {e}")
    finally:
        if 'item' in locals():
            del item


# === Основной цикл мониторинга ===
def monitor_inbox():
    pythoncom.CoInitialize()
    outlook = None
    namespace = None
    inbox = None
    folder = None

    processed_ids = load_processed_ids()
    logging.info(f"Загружено {len(processed_ids)} обработанных писем из файла.")

    try:
        outlook = win32com.client.Dispatch("Outlook.Application")
        namespace = outlook.GetNamespace("MAPI")
        inbox = namespace.GetDefaultFolder(6)
        folder = inbox.Folders[OUTLOOK_FOLDER] if OUTLOOK_FOLDER.lower() != "inbox" else inbox

        logging.info(f"✅ Мониторинг запущен. Режим записи: {WRITE_MODE.upper()}.")
        logging.info(f"📂 Папка: {OUTLOOK_FOLDER}")
        logging.info(f"📬 Ищем письма с темой (без учёта регистра): '{SEARCH_SUBJECT}'")

        time.sleep(5)

        while True:
            try:
                today = datetime.today().date()
                min_date = today - timedelta(days=7)

                messages = folder.Items
                messages.Sort("[ReceivedTime]", True)

                msg_list = []
                logging.info("🔍 Сканируем письма...")

                for msg in messages:
                    try:
                        msg_class = getattr(msg, 'Class', None)
                        if msg_class != 43:
                            continue
                        if not getattr(msg, 'ReceivedTime', None):
                            continue
                        if msg.ReceivedTime.date() < min_date:
                            break
                        msg_list.append(msg)
                    except Exception as e:
                        logging.error(f"Ошибка при анализе письма: {e}")
                        continue

                logging.info(f"📬 Найдено {len(msg_list)} подходящих писем. Начинаем обработку...")

                for msg in msg_list:
                    handle_mail(msg, processed_ids)
                    del msg

                del msg_list

            except Exception as e:
                logging.error(f"❌ Ошибка мониторинга: {e}")

            logging.info("⏳ Ждем 60 секунд до следующей проверки...\n")
            time.sleep(60)

    except Exception as e:
        logging.error(f"❌ Ошибка инициализации Outlook: {e}")
    finally:
        for obj in [folder, inbox, namespace, outlook]:
            if obj:
                del obj
        pythoncom.CoUninitialize()


# === Запуск ===
if __name__ == "__main__":
    try:
        monitor_inbox()
    except Exception as e:
        logging.error(f"❌ Критическая ошибка: {e}")
    finally:
        input("\nНажмите Enter для закрытия...")
